"""
Excel exporter for dataset analysis results.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any
from datetime import datetime
import os


def export_to_excel(analysis_results: Dict[str, Any], output_path: str) -> str:
    """
    Export analysis results to Excel file.
    
    Args:
        analysis_results: Dictionary containing analysis results
        output_path: Output file path
        
    Returns:
        Path to created Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dataset Analysis"
    
    # Add title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "Dataset Analysis Report"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Add metadata
    ws['A2'] = "Generated:"
    ws['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A3'] = "Root Path:"
    ws['B3'] = analysis_results.get('root_path', 'N/A')
    
    # Add summary statistics
    ws['A4'] = "Total Classes:"
    ws['B4'] = analysis_results.get('total_classes', 0)
    ws['A5'] = "Total Annotations:"
    ws['B5'] = analysis_results.get('total_annotations', 0)
    ws['A6'] = "Total Files:"
    ws['B6'] = analysis_results.get('total_files', 0)
    
    # Style metadata section
    for row in range(2, 7):
        ws[f'A{row}'].font = Font(bold=True)
    
    # Add header row for data table
    headers = ['Class Name', 'Annotations', 'Files', 'Locations']
    header_row = 8
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Add data rows
    classes = analysis_results.get('classes', [])
    current_row = header_row + 1
    
    for class_data in classes:
        ws.cell(row=current_row, column=1, value=class_data['class_name'])
        ws.cell(row=current_row, column=2, value=class_data['annotations'])
        ws.cell(row=current_row, column=3, value=class_data['files'])
        
        # Join locations with newlines
        locations = '\n'.join(class_data['locations'])
        ws.cell(row=current_row, column=4, value=locations)
        
        # Style data cells
        for col_idx in range(1, 5):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Center align numbers
            if col_idx in [2, 3]:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        current_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 50
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    
    # Save workbook
    wb.save(output_path)
    return output_path
